"""
builder.py - programmatic Excel workbook generator.

Builds a complete .xlsm workbook from scratch using Base.xlsm as the VBA carrier.
All sheets (Init, column_definitions, part sheets, SBH scaffolds) are generated
from the Python schema in sheet_definitions.py.
"""

import gc
import os
import re
import sys
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import replace

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn, TableFormula
from openpyxl.utils import column_index_from_string, get_column_letter, quote_sheetname

from .sheet_definitions import ColumnDef, SheetDef, TEMPLATE_CONFIGS, ALL_SHEETS, NS_NA

# ── Paths ─────────────────────────────────────────────────────────────────────

_HERE         = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(os.path.dirname(_HERE))
# In a PyInstaller bundle, bundled data is unpacked under sys._MEIPASS (matching
# app.py's resource_path); in dev it is _PROJECT_ROOT (Tool/).
_RESOURCE_ROOT = getattr(sys, "_MEIPASS", _PROJECT_ROOT)
BASE_XLSM     = os.path.join(_RESOURCE_ROOT, "resources", "templates", "Base.xlsm")

# ── Styling constants ─────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1A6468")
HEADER_FONT  = Font(scheme="minor", color="FFFFFF", bold=True,  size=11)
DATA_FONT    = Font(scheme="minor", color="1D1D1F", bold=False, size=11)
BOLD_FONT    = Font(scheme="minor", color="1D1D1F", bold=True,  size=11)
TITLE_FONT   = Font(scheme="major", color="1A6468", bold=True,  size=13)
TAB_COLOR    = "48A9AE"

# Welcome sheet card styling (matches the desktop UI: teal header, near-white
# input fields with a light gray border, system font).
_INPUT_SIDE   = Side(style="thin", color="C7C7CC")
WELCOME_INPUT_FILL   = PatternFill(fill_type="solid", fgColor="FFFFFF")
WELCOME_INPUT_BORDER = Border(left=_INPUT_SIDE, right=_INPUT_SIDE,
                              top=_INPUT_SIDE, bottom=_INPUT_SIDE)
WELCOME_LABEL_FONT   = Font(scheme="minor", color="1D1D1F", bold=True, size=11)
WELCOME_VALUE_FONT   = Font(scheme="minor", color="1D1D1F", bold=False, size=11)
WELCOME_TITLE_FONT   = Font(scheme="minor", color="FFFFFF", bold=True, size=12)
DEFAULT_COL_WIDTH = 25
DATA_VALIDATION_TARGET_MAX_ROW = 5000
DATA_VALIDATION_SOURCE_MAX_ROW = 10000


def _style_header_row(ws):
    ws.row_dimensions[1].height = 20
    for cell in ws[1]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _mark_system_sheet(ws):
    """Apply teal tab color to system/advanced sheets (hidden by default)."""
    ws.sheet_properties.tabColor = TAB_COLOR


def _style_data_rows(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = DATA_FONT


def _set_fixed_col_widths(ws):
    for col_idx in range(1, max(ws.max_column, 1) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_COL_WIDTH


def _style_sheet(ws):
    _style_header_row(ws)
    _style_data_rows(ws)
    _set_fixed_col_widths(ws)


# ── Init sheet constants ──────────────────────────────────────────────────────

# Column order in Init data rows (after the Sheet Name index column)
_INIT_COLS = [
    "Sheet Name",
    "Convert",
    "Has Collections",
    "Collect Cols",
    "# of Collect Rows",
    "Has Descripts",
    "Descript Cols",
    "Descript Start Row",
    "Lib Start Row",
    "SBOL Object Type",
    "Molecule Type",
    "Role",
]
_INIT_HEADER_ROW = 10

# column_definitions rows for the Init sheet itself - maps Init metadata columns
# to SBOL terms so the compiler can inject them into each converted sheet.
_INIT_CD_ROWS = [
    dict(sheet_name="Init", col_name="SBOL Object Type", sbol_term="sbol_objectType",
         namespace="http://sbols.org/v2#", col_type="String"),
    dict(sheet_name="Init", col_name="Molecule Type", sbol_term="sbol_types",
         namespace="http://sbols.org/v2#", col_type="String",
         sheet_lk=True, onto_name="BIOPAX", lk_sheet="molecule_types",
         from_col="A", to_col="B"),
    dict(sheet_name="Init", col_name="Role", sbol_term="sbol_roles",
         namespace="http://sbols.org/v2#", col_type="URI",
         sheet_lk=True, onto_name="SO", lk_sheet="ontology_terms",
         from_col="A", to_col="B"),
]

# Reference lookup sheets that must appear in Init with Convert=False so the
# compiler builds compiled_sheets entries for them - required for Sheet Lookup
# resolution at conversion time.
_REFERENCE_SHEETS = ("ontology_terms", "organism_terms", "molecule_types")

# All reference sheets carried from Base.xlsm that are pruned when the selected
# sheets do not use them. nci_thesaurus is never referenced by the current
# schema, so it is always pruned.
_ALL_REFERENCE_SHEETS = ("ontology_terms", "organism_terms", "molecule_types", "nci_thesaurus")


# Resolved ontology URIs for the role labels used by SheetDefs. Lets templates
# that do not otherwise need ontology_terms (no column looks it up) write the role
# URI straight into Init and drop the ontology_terms sheet entirely.
_ROLE_URI = {
    "promoter":            "http://identifiers.org/so/SO:0000167",
    "ribosome_entry_site": "http://identifiers.org/so/SO:0000139",
    "CDS":                 "http://identifiers.org/so/SO:0000316",
    "terminator":          "http://identifiers.org/so/SO:0000141",
    "RNA":                 "http://identifiers.org/so/SO:0000356",
    "signal":              "http://identifiers.org/ncit/NCIT:C43382",
    # NCIT:C14419 "Organism Strain" - the host organism, used by chassis.
    "organism_strain":     "http://identifiers.org/ncit/NCIT:C14419",
    "medium":              "http://identifiers.org/ncit/NCIT:C48164",
    # NCIT:C97158 "Genetically Modified Organism" - an engineered strain, which
    # is a distinct concept from the unmodified host above.
    "genetically_modified_organism": "http://purl.obolibrary.org/obo/NCIT_C97158",
    "sample_design":       "https://wiki.synbiohub.org/wiki/Terms/SynBioSuite#SampleDesign",
    "supplement":          "http://purl.obolibrary.org/obo/PATO_0000033",
}


def _needed_reference_sheets(sheets: list) -> set:
    """Reference sheets needed by column lookups or Init Molecule Type injection.

    molecule_types - any sheet with a molecule_type, or any column that looks it up.
    ontology_terms / organism_terms / nci_thesaurus - only if a column looks them up.
    The role-driven need for ontology_terms is decided in generate(): roles can be
    written as direct URIs (see _ROLE_URI), which removes that dependency.
    """
    needed = set()
    if any(getattr(s, "molecule_type", "") for s in sheets):
        needed.add("molecule_types")
    for s in sheets:
        for col in s.columns:
            lk = getattr(col, "lookup_sheet", None)
            if lk in _ALL_REFERENCE_SHEETS:
                needed.add(lk)
    return needed

# column_definitions header row (written as row 1)
_CD_COLS = [
    "Sheet Name",
    "Column Name",
    "SBOL Term",
    "Namespace URL",
    "Type",
    "Split On",
    "Pattern",
    "Multicolumn",
    "Tyto Lookup",
    "Sheet Lookup",
    "Replacement Lookup",
    "Object_ID Lookup",
    "Parent Lookup",
    "Ontology Name",
    "Lookup Sheet Name",
    "From Col",
    "To Col",
    "comma_append",
    "translate_target",
]

# Maps lookup_sheet name → column A header text (for _DropdownMap Source Header)
_LOOKUP_DISPLAY_HEADER = {
    "ontology_terms":          "Role",
    "organism_terms":          "Organism",
    "molecule_types":          "Name",
    "nci_thesaurus":           "NAME",
    "SBH_chassis_collections": "Name",
    "SBH_plasmids_collections": "Name",
    "SBH_chemicals_collection": "Name",
    "SBH_strains_collection":   "Name",
    "SBH_media_collection":     "Name",
    "SBH_supplements_collection": "Name",
    "SBH_sampledesigns_collection": "Name",
}

# Single-value object references get a native dropdown sourced from the ID column
# of the sheet that holds the referenced objects (a cds encodes a protein, an
# assay belongs to a study, a sample to an assay). Multi-value reference columns
# (comma_append, or split-on-comma like sample design Supplements) use the VBA
# picker instead, since a native list validation can hold only one value.
_OBJECT_REF_DROPDOWN = {
    ("cds", "Encodes for"):   ("protein", "Protein ID"),
    ("ncrna", "Encodes for"): ("rna", "RNA ID"),
    ("assay", "Study ID"):    ("study", "Study ID"),
    ("sample", "Assay ID"):   ("assay", "Assay ID"),
}


# ── Main entry point ──────────────────────────────────────────────────────────

def generate(config: dict, progress_cb=None) -> str:
    """
    Generate a complete Excel template workbook.

    config keys:
        template_type  : str - key in TEMPLATE_CONFIGS, or "custom"
        custom_sheets  : list[str] - sheet names from ALL_SHEETS (for custom type)
        output_folder  : str - directory to save the generated file
        metadata       : dict with keys:
            library_name, collection_id, version, author, email, lab, institution,
            description, pubmed_id, sbol_version, domain, master_collection

    Returns the path to the saved file.
    """
    def _progress(msg: str):
        if progress_cb:
            progress_cb(msg)

    template_type = config["template_type"]
    metadata      = config.get("metadata", {})
    output_folder = config["output_folder"]

    library_name = metadata.get("library_name", "MyLibrary").strip() or "MyLibrary"
    sbol_version = int(metadata.get("sbol_version", 2))

    safe_name    = _safe_filename(library_name)
    type_label   = _type_label(template_type, config)
    out_filename = f"{safe_name}_{type_label}.xlsm"
    out_path     = os.path.join(output_folder, out_filename)

    if not os.path.exists(BASE_XLSM):
        raise FileNotFoundError(f"Base workbook not found: {BASE_XLSM}")

    # 1. Select sheet list
    _progress("Resolving sheet list...")
    sheets = _resolve_sheets(template_type, config)

    # Append user-defined custom sheets from the UI modal editor
    for raw_sheet in config.get("user_custom_sheets", []):
        sheets.append(_ui_sheet_to_sheetdef(raw_sheet))

    # order the data-sheet tabs by the user's chosen arrangement, if given.
    # `sheet_order` is an ordered list of sheet names (built-in + custom); any
    # sheet not named falls to the end in its existing relative order.
    order = config.get("sheet_order")
    if order:
        pos = {name: i for i, name in enumerate(order)}
        sheets.sort(key=lambda s: pos.get(s.name, len(pos)))

    # F4: apply per-sheet column reordering chosen in the UI. SheetDefs are shared
    # module-level objects, so replace with a copy rather than mutating in place.
    # Reorder-only: every original column is preserved (columns resolve by name
    # everywhere, so position is purely cosmetic).
    col_orders = config.get("column_orders") or {}
    if col_orders:
        sheets = [replace(s, columns=_reorder_columns(s.columns, col_orders[s.name]))
                  if col_orders.get(s.name) else s
                  for s in sheets]

    # 2. Copy base workbook (preserves VBA)
    _progress("Copying base workbook...")
    shutil.copy2(BASE_XLSM, out_path)
    wb = openpyxl.load_workbook(out_path, keep_vba=True)

    # 2b. Prune reference sheets the selected sheets do not use. Sheet-level roles
    # are written as direct URIs whenever every role is known, independently of
    # whether ontology_terms survives for some column's dropdown. Keeping these two
    # decisions separate means ontology_terms only has to carry the terms a column
    # actually offers, not the ones the Init roles happen to need.
    needed_refs = _needed_reference_sheets(sheets)
    roles_present = {s.role for s in sheets if getattr(s, "role", "")}
    direct_uri_roles = (
        bool(roles_present)
        and roles_present <= set(_ROLE_URI)                 # every role has a known URI
    )
    if roles_present and not direct_uri_roles:
        # an unknown role still needs the label + ontology_terms lookup
        needed_refs = needed_refs | {"ontology_terms"}
    for ref in _ALL_REFERENCE_SHEETS:
        if ref not in needed_refs and ref in wb.sheetnames:
            del wb[ref]

    # 3. Write welcome sheet metadata
    _progress("Writing welcome metadata...")
    _write_welcome(wb, metadata, library_name, type_label)

    # 4. Create Init sheet
    _progress("Building Init sheet...")
    _build_init(wb, sheets, sbol_version, library_name, needed_refs, direct_uri_roles)

    # 5. Create column_definitions sheet
    _progress("Building column_definitions sheet...")
    _build_column_definitions(wb, sheets, direct_uri_roles, needed_refs)

    # 6. Create each selected sheet
    for sheet_def in sheets:
        _progress(f"Building sheet: {sheet_def.name}...")
        _build_sheet(wb, sheet_def)

    # 7. Create SBH scaffold sheets
    _progress("Creating SBH scaffold sheets...")
    _build_sbh_scaffolds(wb, sheets)

    # 8. Apply native Excel dropdowns for sheet-backed lookup columns
    _progress("Applying data validation dropdowns...")
    _apply_data_validations(wb, sheets)

    # 10. Filter flapjack_cols to selected sheets
    _progress("Filtering flapjack data...")
    _filter_flapjack_cols(wb, sheets)

    # 11. Style reference sheets (already styled in Base.xlsm; re-apply tab color)
    _apply_reference_styling(wb)

    # 12. Standardize column widths across generated and inherited base sheets
    _apply_workbook_col_widths(wb)

    _apply_sheet_visibility(wb)

    # 13. Set welcome as the active sheet on open
    if "welcome" in wb.sheetnames:
        wb.active = wb.index(wb["welcome"])

    # Open large: Excel has no "maximized" flag in the file, but a big saved
    # window size (in twips) makes the workbook open near-full-screen without a
    # macro. Excel clamps to the actual screen. Best-effort; not a true maximize.
    if wb.views:
        wb.views[0].windowWidth = 32767
        wb.views[0].windowHeight = 32767
        wb.views[0].xWindow = 0
        wb.views[0].yWindow = 0

    # 14. Save
    _progress("Saving workbook...")
    wb.save(out_path)

    # 15. Inject native Excel checkbox controls. openpyxl cannot write this
    # feature and strips it on save, so it is applied as a post-save OOXML step.
    _progress("Applying native checkboxes...")
    _apply_native_checkboxes(out_path)

    _progress(f"Done: {out_path}")
    return out_path


# ── Step helpers ──────────────────────────────────────────────────────────────

def _reorder_columns(columns: list, order: list) -> list:
    """Return `columns` reordered to match the names in `order`.

    Reorder-only and lossless: columns named in `order` come first in that order,
    any columns not named keep their original relative order at the end, and every
    original column is preserved exactly once. Unknown names in `order` are ignored.
    """
    by_name = {c.name: c for c in columns}
    result, seen = [], set()
    for name in order:
        c = by_name.get(name)
        if c is not None and name not in seen:
            result.append(c)
            seen.add(name)
    for c in columns:
        if c.name not in seen:
            result.append(c)
            seen.add(c.name)
    return result


def _resolve_sheets(template_type: str, config: dict) -> list:
    if template_type == "custom":
        # User picked from the full catalog - preserve catalog order
        selected = set(config.get("selected_sheets", []))
        return [s for s in ALL_SHEETS.values() if s.name in selected]
    full_list = list(TEMPLATE_CONFIGS.get(template_type, []))
    selected = config.get("selected_sheets")
    if selected is not None:
        selected_set = set(selected)
        full_list = [s for s in full_list if s.name in selected_set]
    return full_list


def _ui_sheet_to_sheetdef(sheet_data: dict) -> SheetDef:
    """Convert a user-defined sheet dict (from the UI modal editor) into a SheetDef."""
    cols = []
    for c in sheet_data.get("columns", []):
        header = c.get("header", "").strip()
        if not header:
            continue
        sbol_term = c.get("sbolTerm", "Not_applicable") or "Not_applicable"
        col_type  = c.get("type", "String") or "String"
        namespace = c.get("namespace", NS_NA) or NS_NA

        col = ColumnDef(
            name=header,
            tooltip=c.get("tooltip", ""),
            sbol_term=sbol_term,
            namespace=namespace,
            col_type=col_type,
            tyto_lookup=bool(c.get("tytoLookup")),
            sheet_lookup=bool(c.get("sheetLookup")),
            object_id_lookup=bool(c.get("objectIdLookup")),
            lookup_sheet=c.get("lookupSheet") or None,
            from_col=c.get("fromCol") or None,
            to_col=c.get("toCol") or None,
            ontology_name=c.get("ontologyName") or None,
        )
        cols.append(col)

    sheet_name = sheet_data.get("name", "custom").strip().lower()
    return SheetDef(
        name=sheet_name,
        display_name=sheet_data.get("displayName", sheet_name.title()),
        sbol_object_type=sheet_data.get("sbolObjectType", ""),
        molecule_type=sheet_data.get("moleculeType", ""),
        role=sheet_data.get("role", ""),
        flapjack_object=sheet_data.get("flapjackObject") or None,
        sbh_collections=[],
        columns=cols,
        name_column=sheet_data.get("nameColumn") or None,
    )


def _write_welcome(wb, metadata: dict, library_name: str, type_label: str):
    ws = wb["welcome"]
    _write_welcome_field(ws, "Author",         metadata.get("author", ""))
    _write_welcome_field(ws, "Email",          metadata.get("email", ""))
    _write_welcome_field(ws, "Laboratory",     metadata.get("lab", ""))
    _write_welcome_field(ws, "Institution",    metadata.get("institution", ""))
    _write_welcome_field(ws, "Name",           library_name)
    _write_welcome_field(ws, "Description",    metadata.get("description", ""))
    _write_welcome_field(ws, "Version",        metadata.get("version", 1))
    _write_welcome_field(ws, "PubMed IDs",     metadata.get("pubmed_id", ""))
    _write_welcome_field(ws, "Domain",         metadata.get("domain", "").rstrip("/"))
    _write_welcome_field(ws, "Master Collection", metadata.get("master_collection", ""))
    _write_welcome_field(ws, "Template Type",  type_label)
    # ID: an explicit value if the wizard supplied one, otherwise a formula that
    # derives it from the Name cell exactly as the data-sheet ID auto-fill does.
    _write_welcome_id(ws, metadata.get("collection_id", ""))
    _style_welcome_sheet(ws)


def _welcome_label_row(ws, label: str):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip() == label:
            return r
    return None


def _write_welcome_id(ws, explicit_id: str):
    id_row = _welcome_label_row(ws, "ID")
    if id_row is None:
        return
    if explicit_id:
        ws.cell(id_row, 3).value = explicit_id
        return
    name_row = _welcome_label_row(ws, "Name")
    if name_row is None:
        return
    name_ref = f"$C${name_row}"
    ws.cell(id_row, 3).value = (
        f'=IF({name_ref}="","",'
        f'SUBSTITUTE(SUBSTITUTE(TRIM({name_ref})," ","_"),"-","_"))'
    )


# Boolean welcome fields rendered as native checkboxes (their value cell in col C).
_WELCOME_CHECKBOX_LABELS = ("Show Common Sheets", "Show Advanced Sheets")


def _style_welcome_sheet(ws):
    """Give the welcome sheet the desktop UI's card look: a teal title bar, bold
    labels, and near-white input-field value cells. Every value cell is styled
    the same way (no read-only distinction) for consistency; the two boolean
    rows are turned into checkboxes later by the native-checkbox pass."""
    ws.sheet_view.showGridLines = False

    # 1. wipe stray borders/fills left in the base template (the "random lines")
    for row in ws.iter_rows(min_row=1, max_row=45, min_col=1, max_col=14):
        for c in row:
            c.border = Border()
            if c.fill is not None and c.fill.patternType:
                c.fill = PatternFill()

    label_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value]
    if not label_rows:
        return
    top, bottom = min(label_rows), max(label_rows)

    # 2. teal title bar one row above the fields
    title_row = top - 1 if top > 1 else 1
    for col in (2, 3):
        cell = ws.cell(title_row, col)
        cell.fill = HEADER_FILL
        cell.font = WELCOME_TITLE_FONT
        cell.alignment = Alignment(vertical="center")
    ws.cell(title_row, 2).value = "Library Information"
    ws.row_dimensions[title_row].height = 22

    # 3. labels bold; value cells styled as input fields, except the boolean
    #    option rows (they become checkboxes) and the blank spacer that sets the
    #    options apart as a group rather than more data to enter.
    for r in range(top, bottom + 1):
        label = str(ws.cell(r, 2).value or "").strip()
        if not label:
            continue                       # spacer row stays blank/white
        lbl = ws.cell(r, 2)
        lbl.font = WELCOME_LABEL_FONT
        lbl.alignment = Alignment(vertical="center")
        val = ws.cell(r, 3)
        val.font = WELCOME_VALUE_FONT
        val.alignment = Alignment(vertical="center", horizontal="left")
        if label not in _WELCOME_CHECKBOX_LABELS:
            val.fill = WELCOME_INPUT_FILL
            val.border = WELCOME_INPUT_BORDER


def _write_welcome_field(ws, label: str, value):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip() == label:
            ws.cell(r, 3).value = value
            return


def _build_init(wb, sheets: list, sbol_version: int, library_name: str,
                needed_refs=None, direct_uri_roles=False):
    if needed_refs is None:
        needed_refs = set(_REFERENCE_SHEETS)
    if "Init" in wb.sheetnames:
        del wb["Init"]

    ws = wb.create_sheet("Init")

    # Metadata rows (rows 1-9)
    ws.cell(1, 1).value = "SBOL Version"
    ws.cell(1, 2).value = sbol_version
    ws.cell(2, 1).value = "Library Name"
    ws.cell(2, 2).value = library_name

    # Header row (row 10)
    for c, hdr in enumerate(_INIT_COLS, 1):
        ws.cell(_INIT_HEADER_ROW, c).value = hdr

    # Data rows (row 11+)
    col_map = {h: i for i, h in enumerate(_INIT_COLS, 1)}
    row = _INIT_HEADER_ROW + 1
    for sdef in sheets:
        # Only sheets with a real SBOL object type can be converted to SBOL
        # objects. Flapjack-only sheets (empty sbol_object_type, e.g. measurement)
        # are written with Convert=False so their data is still loaded for
        # lookups/Flapjack but they never enter parse_objects(), which would
        # crash on the missing sbol_objectType column otherwise.
        convert = bool(sdef.sbol_object_type and str(sdef.sbol_object_type).strip())
        ws.cell(row, col_map["Sheet Name"]).value      = sdef.name
        ws.cell(row, col_map["Convert"]).value         = convert
        ws.cell(row, col_map["Has Collections"]).value = False
        ws.cell(row, col_map["Collect Cols"]).value    = 0
        ws.cell(row, col_map["# of Collect Rows"]).value = 0
        ws.cell(row, col_map["Has Descripts"]).value   = False
        ws.cell(row, col_map["Descript Cols"]).value   = 0
        ws.cell(row, col_map["Descript Start Row"]).value = 0
        ws.cell(row, col_map["Lib Start Row"]).value   = 1
        ws.cell(row, col_map["SBOL Object Type"]).value = sdef.sbol_object_type
        ws.cell(row, col_map["Molecule Type"]).value   = sdef.molecule_type
        # write the role URI directly when ontology_terms is being dropped;
        # otherwise the label (resolved via the ontology_terms lookup at convert).
        role_val = sdef.role
        if direct_uri_roles and role_val:
            role_val = _ROLE_URI.get(role_val, role_val)
        ws.cell(row, col_map["Role"]).value            = role_val
        row += 1

    for ref_name in _REFERENCE_SHEETS:
        if ref_name not in needed_refs:  # only list reference sheets in use
            continue
        ws.cell(row, col_map["Sheet Name"]).value        = ref_name
        ws.cell(row, col_map["Convert"]).value           = False
        ws.cell(row, col_map["Has Collections"]).value   = False
        ws.cell(row, col_map["Collect Cols"]).value      = 0
        ws.cell(row, col_map["# of Collect Rows"]).value = 0
        ws.cell(row, col_map["Has Descripts"]).value     = False
        ws.cell(row, col_map["Descript Cols"]).value     = 0
        ws.cell(row, col_map["Descript Start Row"]).value = 0
        ws.cell(row, col_map["Lib Start Row"]).value     = 0
        ws.cell(row, col_map["SBOL Object Type"]).value  = ""
        ws.cell(row, col_map["Molecule Type"]).value     = ""
        ws.cell(row, col_map["Role"]).value              = ""
        row += 1

    seen_scaffolds = set()
    for sdef in sheets:
        for scaffold_name in sdef.sbh_collections:
            if scaffold_name in seen_scaffolds:
                continue
            seen_scaffolds.add(scaffold_name)
            ws.cell(row, col_map["Sheet Name"]).value        = scaffold_name
            ws.cell(row, col_map["Convert"]).value           = False
            ws.cell(row, col_map["Has Collections"]).value   = False
            ws.cell(row, col_map["Collect Cols"]).value      = 0
            ws.cell(row, col_map["# of Collect Rows"]).value = 0
            ws.cell(row, col_map["Has Descripts"]).value     = False
            ws.cell(row, col_map["Descript Cols"]).value     = 0
            ws.cell(row, col_map["Descript Start Row"]).value = 0
            ws.cell(row, col_map["Lib Start Row"]).value     = 0
            ws.cell(row, col_map["SBOL Object Type"]).value  = ""
            ws.cell(row, col_map["Molecule Type"]).value     = ""
            ws.cell(row, col_map["Role"]).value              = ""
            row += 1

    _style_sheet(ws)
    _mark_system_sheet(ws)


def _build_column_definitions(wb, sheets: list, direct_uri_roles=False,
                              needed_refs=None):
    if "column_definitions" in wb.sheetnames:
        del wb["column_definitions"]

    ws = wb.create_sheet("column_definitions")

    # Header row
    for c, hdr in enumerate(_CD_COLS, 1):
        ws.cell(1, c).value = hdr

    col_map = {h: i for i, h in enumerate(_CD_COLS, 1)}
    row = 2

    # Init metadata rows (so compiler can map SBOL Object Type, Molecule Type, Role)
    for row_kwargs in _INIT_CD_ROWS:
        # when roles are written as direct URIs, the Role row must NOT be an
        # ontology_terms sheet lookup (that sheet is pruned) - pass the URI through.
        if direct_uri_roles and row_kwargs.get("col_name") == "Role":
            row_kwargs = dict(sheet_name="Init", col_name="Role",
                              sbol_term="sbol_roles",
                              namespace="http://sbols.org/v2#", col_type="URI")
        # Same for Molecule Type: molecule_types is pruned when no sheet declares
        # one, so pointing a lookup at it would leave a dangling reference.
        if (row_kwargs.get("col_name") == "Molecule Type"
                and needed_refs is not None
                and "molecule_types" not in needed_refs):
            row_kwargs = dict(sheet_name="Init", col_name="Molecule Type",
                              sbol_term="sbol_types",
                              namespace="http://sbols.org/v2#", col_type="String")
        _write_cd_row(ws, col_map, row, **row_kwargs)
        row += 1

    # One row per column per sheet (excluding VBA-only columns from converter processing)
    for sdef in sheets:
        for col in sdef.columns:
            if col.sbol_term == "Not_applicable" and col.translate_target is None:
                # Only write to column_definitions if it has a real SBOL term
                # or is a translate trigger (VBA reads translate_target from col_defs)
                if col.col_type == "Not_applicable":
                    # Still write so VBA can read comma_append / translate_target
                    pass
            _write_cd_col(ws, col_map, row, sdef.name, col)
            row += 1

    _style_sheet(ws)
    _mark_system_sheet(ws)


def _write_cd_row(ws, col_map, row, sheet_name, col_name, sbol_term,
                  namespace, col_type, split_on='""', pattern=None,
                  tyto=False, sheet_lk=False, repl_lk=False,
                  obj_id_lk=False, par_lk=False, onto_name=None,
                  lk_sheet=None, from_col=None, to_col=None,
                  comma_append=False, translate_target=None):
    ws.cell(row, col_map["Sheet Name"]).value     = sheet_name
    ws.cell(row, col_map["Column Name"]).value    = col_name
    ws.cell(row, col_map["SBOL Term"]).value      = sbol_term
    ws.cell(row, col_map["Namespace URL"]).value  = namespace
    ws.cell(row, col_map["Type"]).value           = col_type
    ws.cell(row, col_map["Split On"]).value       = split_on
    ws.cell(row, col_map["Pattern"]).value        = pattern or '""'
    ws.cell(row, col_map["Multicolumn"]).value    = ""
    ws.cell(row, col_map["Tyto Lookup"]).value    = tyto
    ws.cell(row, col_map["Sheet Lookup"]).value   = sheet_lk
    ws.cell(row, col_map["Replacement Lookup"]).value = repl_lk
    ws.cell(row, col_map["Object_ID Lookup"]).value   = obj_id_lk
    ws.cell(row, col_map["Parent Lookup"]).value  = par_lk
    ws.cell(row, col_map["Ontology Name"]).value  = onto_name or ""
    ws.cell(row, col_map["Lookup Sheet Name"]).value  = lk_sheet or ""
    ws.cell(row, col_map["From Col"]).value       = from_col or ""
    ws.cell(row, col_map["To Col"]).value         = to_col or ""
    ws.cell(row, col_map["comma_append"]).value   = comma_append
    ws.cell(row, col_map["translate_target"]).value = translate_target or ""


def _write_cd_col(ws, col_map, row, sheet_name: str, col: ColumnDef):
    _write_cd_row(
        ws, col_map, row,
        sheet_name    = sheet_name,
        col_name      = col.name,
        sbol_term     = col.sbol_term,
        namespace     = col.namespace,
        col_type      = col.col_type,
        split_on      = col.split_on,
        pattern       = col.pattern,
        tyto          = col.tyto_lookup,
        sheet_lk      = col.sheet_lookup,
        repl_lk       = col.replacement_lookup,
        obj_id_lk     = col.object_id_lookup,
        par_lk        = col.parent_lookup,
        onto_name     = col.ontology_name,
        lk_sheet      = col.lookup_sheet,
        from_col      = col.from_col,
        to_col        = col.to_col,
        comma_append      = col.comma_append,
        translate_target  = col.translate_target,
    )


def _build_sheet(wb, sdef: SheetDef):
    if sdef.name in wb.sheetnames:
        del wb[sdef.name]

    ws = wb.create_sheet(sdef.name)

    # Row 1: headers with tooltip comments
    for c, col in enumerate(sdef.columns, 1):
        cell = ws.cell(1, c)
        cell.value = col.name
        comment = Comment(text=col.tooltip, author="Excel2SBOL")
        comment.width  = 320
        comment.height = 90
        cell.comment   = comment

    # Row 2: formula row (returns the columns that are Excel calculated columns)
    calc_formulas: dict = {}
    if len(sdef.columns) >= 2:
        calc_formulas = _write_formula_row(ws, sdef)

    # Excel Table wrapping rows 1+
    if sdef.columns:
        table_name = re.sub(r"\W+", "_", sdef.name).strip("_") or "Sheet"
        last_col   = get_column_letter(len(sdef.columns))
        # Build columns explicitly so the ID/length columns carry a
        # calculatedColumnFormula -> Excel auto-fills them on new rows.
        table_columns = []
        for i, col in enumerate(sdef.columns, 1):
            tc = TableColumn(id=i, name=col.name)
            if col.name in calc_formulas:
                tc.calculatedColumnFormula = TableFormula(attr_text=calc_formulas[col.name])
            table_columns.append(tc)
        tbl = Table(
            displayName  = table_name,
            ref          = f"A1:{last_col}2",
            tableColumns = table_columns,
        )
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True,  showColumnStripes=False,
        )
        ws.add_table(tbl)

    _style_sheet(ws)


def _write_formula_row(ws, sdef: SheetDef) -> dict:
    """Write ID auto-fill, length, and Translate values in row 2.

    Uses fully-qualified structured references (TableName[[#This Row],[Col]])
    which match native Excel format and survive column reordering. Shorthand
    [@[Col]] syntax is rejected by Excel when written by third-party tools, so
    the full form is required.

    Returns {header: formula-without-leading-'='} for the columns that should be
    Excel *calculated columns* (ID auto-fill, length), so the caller can mark the
    matching TableColumn with a calculatedColumnFormula and Excel fills them down
    as rows are added. The Translate columns hold a static Unicode ballot box
    toggled by VBA, so they are deliberately NOT calculated columns.
    """
    headers  = {col.name: i for i, col in enumerate(sdef.columns, 1)}
    tbl_name = re.sub(r"\W+", "_", sdef.name).strip("_") or "Sheet"
    calc: dict = {}

    def _sr(col_name):
        return f"{tbl_name}[[#This Row],[{col_name}]]"

    # ID auto-fill
    dn       = sdef.display_name
    name_hdr = f"{dn} Name"
    id_hdr   = f"{dn} ID"
    if name_hdr in headers and id_hdr in headers:
        # Case is preserved so the derived ID matches what the user typed as the
        # name; reference columns resolve against the ID, so lowercasing here made
        # a part named GFP unreachable as GFP.
        f = (f'IF({_sr(name_hdr)}="","",'
             f'SUBSTITUTE(SUBSTITUTE(TRIM({_sr(name_hdr)})," ","_"),"-","_"))')
        ws.cell(2, headers[id_hdr]).value = "=" + f
        calc[id_hdr] = f

    # Sequence length
    seq_col = headers.get("Sequence")
    if seq_col:
        for length_name in ("Length (bp)", "Length (nt)", "Length (aa)"):
            len_col = headers.get(length_name)
            if len_col:
                f = (f'IF({_sr("Sequence")}="","",'
                     f'LEN(TRIM(CLEAN(SUBSTITUTE({_sr("Sequence")}," ","")))))')
                ws.cell(2, len_col).value = "=" + f
                calc[length_name] = f
                break

    # Translate trigger columns show a Unicode ballot box (U+2610, empty) that the
    # VBA double-click handler toggles to a checked box (U+2611). A native 2024
    # checkbox cell-control was abandoned because it does not propagate on Excel
    # Table expand and Excel will not carry it through a style, auto-fill, or
    # dataDxfId; a glyph is plain text, so it fills and toggles on new rows with
    # no clipboard. Centered so it reads like a checkbox. NOT a calculated column.
    for col in sdef.columns:
        if col.translate_target and col.name in headers:
            c = ws.cell(2, headers[col.name])
            c.value = "☐"
            c.alignment = Alignment(horizontal="center", vertical="center")

    return calc


def _build_sbh_scaffolds(wb, sheets: list):
    created = set()
    for sdef in sheets:
        for scaffold_name in sdef.sbh_collections:
            if scaffold_name in created or scaffold_name in wb.sheetnames:
                continue
            ws = wb.create_sheet(scaffold_name)
            ws.cell(1, 1).value = "Name"
            ws.cell(1, 2).value = "URI"
            _style_sheet(ws)
            _mark_system_sheet(ws)
            # Wrap in an Excel Table so its dropdown source can be a structured
            # reference that auto-grows as the user pastes lookup values, with no
            # trailing blanks (unlike a fixed range).
            tbl_name = re.sub(r"\W+", "_", scaffold_name).strip("_") or "SBH"
            table = Table(displayName=tbl_name, ref="A1:B2")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(table)
            created.add(scaffold_name)


# Columns whose values are booleans but which carry no lookup sheet, so they get
# a literal TRUE/FALSE list validation instead of a sourced dropdown.
_BOOL_LIST_COLUMNS = ("Degrades",)


def _apply_data_validations(wb, sheets: list):
    """Apply native Excel list dropdowns for sheet-backed lookup columns."""
    for sdef in sheets:
        if sdef.name not in wb.sheetnames:
            continue

        target_ws = wb[sdef.name]
        target_headers = {
            str(target_ws.cell(1, c).value or ""): c
            for c in range(1, target_ws.max_column + 1)
        }

        # Literal TRUE/FALSE dropdown for boolean-valued columns that are not
        # backed by a lookup sheet.
        for col in sdef.columns:
            if col.name not in _BOOL_LIST_COLUMNS or col.name not in target_headers:
                continue
            target_col = get_column_letter(target_headers[col.name])
            target_range = f"{target_col}2:{target_col}{max(2, target_ws.max_row)}"
            validation = DataValidation(type="list", formula1='"TRUE,FALSE"',
                                        allow_blank=True)
            target_ws.add_data_validation(validation)
            validation.add(target_range)

        for col in sdef.columns:
            if not (col.sheet_lookup and col.lookup_sheet):
                continue
            if col.comma_append:
                # Multi-value column. A native list validation holds one value at
                # a time, so pairing it with the MultiSelectPicker would put two
                # competing dropdowns on the same cell; the picker owns these.
                continue
            if col.name not in target_headers or col.lookup_sheet not in wb.sheetnames:
                continue

            source_ws = wb[col.lookup_sheet]
            source_col_idx = _lookup_source_col_index(source_ws, col)
            if source_col_idx is None:
                continue

            # Table structured reference when the source is a table (SBH
            # scaffolds), exact data extent otherwise (fixed reference
            # vocabularies). Both avoid the trailing blanks a fixed 10000-row
            # range produces.
            source_header = str(source_ws.cell(1, source_col_idx).value or "")
            formula = _object_ref_source(wb, source_ws, source_header, source_col_idx)
            if formula is None:
                continue

            target_col = get_column_letter(target_headers[col.name])
            # Apply to the table's single data row only; Excel extends the
            # validation to new rows as the table grows, so the dropdown does not
            # litter hundreds of empty rows below the table.
            target_range = f"{target_col}2:{target_col}{max(2, target_ws.max_row)}"
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            target_ws.add_data_validation(validation)
            validation.add(target_range)

        # Native dropdown for single-value object references (e.g. cds Encodes
        # for -> proteins), sourced from the referenced sheet's ID column.
        for col in sdef.columns:
            key = (sdef.name, col.name)
            if key not in _OBJECT_REF_DROPDOWN or col.name not in target_headers:
                continue
            src_sheet, src_id_header = _OBJECT_REF_DROPDOWN[key]
            if src_sheet not in wb.sheetnames:
                continue
            src_ws = wb[src_sheet]
            src_headers = {str(src_ws.cell(1, c).value or ""): c
                           for c in range(1, src_ws.max_column + 1)}
            if src_id_header not in src_headers:
                continue
            # Source from the referenced sheet's TABLE column via a defined name,
            # not a fixed row range: a big fixed range would show the empty
            # formula cells (the ID auto-fill sits in every table row) as blank
            # dropdown options. A table structured reference auto-sizes to the
            # real rows and grows as objects are added.
            formula = _object_ref_source(wb, src_ws, src_id_header,
                                         src_headers[src_id_header])
            if formula is None:
                continue
            target_col = get_column_letter(target_headers[col.name])
            # Apply to the table's single data row only; Excel extends the
            # validation to new rows as the table grows, so the dropdown does not
            # litter hundreds of empty rows below the table.
            target_range = f"{target_col}2:{target_col}{max(2, target_ws.max_row)}"
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            target_ws.add_data_validation(validation)
            validation.add(target_range)


def _object_ref_source(wb, src_ws, id_header, id_col_idx):
    """A data-validation source for an object-reference dropdown.

    Prefers a workbook defined name that resolves to the source sheet's table
    column (`table[Header]`), which auto-sizes to the real rows and avoids the
    blank options a fixed range produces (the ID column is formula-filled in
    every table row). Falls back to a bounded range if the sheet has no table.
    Returns the validation formula string, or None if it cannot be built.
    """
    # Each generated data sheet carries exactly one Excel Table.
    table_names = list(getattr(src_ws, "tables", {}) or {})
    table_name = table_names[0] if table_names else None

    if table_name:
        safe = re.sub(r"\W+", "_", f"{table_name}_{id_header}").strip("_")
        dn_name = f"dv_{safe}"
        if dn_name not in wb.defined_names:
            from openpyxl.workbook.defined_name import DefinedName
            wb.defined_names.append(
                DefinedName(name=dn_name, attr_text=f"{table_name}[{id_header}]"))
        return f"={dn_name}"

    col = get_column_letter(id_col_idx)
    last = max(src_ws.max_row, 2)
    return f"{quote_sheetname(src_ws.title)}!${col}$2:${col}${last}"


def _lookup_source_col_index(source_ws, col: ColumnDef):
    if col.from_col:
        try:
            return column_index_from_string(str(col.from_col))
        except ValueError:
            return None

    source_header = _LOOKUP_DISPLAY_HEADER.get(col.lookup_sheet)
    if source_header:
        for c in range(1, source_ws.max_column + 1):
            if str(source_ws.cell(1, c).value or "") == source_header:
                return c

    return 1


def _filter_flapjack_cols(wb, sheets: list):
    fj_sheet = "flapjack_cols"
    if fj_sheet not in wb.sheetnames:
        return

    ws = wb[fj_sheet]
    selected_names = {sdef.name for sdef in sheets}

    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        sheet_name = ws.cell(r, 1).value
        if sheet_name is not None and sheet_name not in selected_names:
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    # Nothing left but the header means no selected sheet maps to Flapjack, so
    # the sheet would only be dead weight in the generated workbook.
    if not any(ws.cell(r, 1).value for r in range(2, ws.max_row + 1)):
        del wb[fj_sheet]


# Sheets the user never edits by hand. Hidden at generation time rather than
# left to the SheetVisibility macro, which only runs once the workbook is opened
# in Excel with macros enabled.
_HIDDEN_SYSTEM_SHEETS = ("Init", "column_definitions",
                         # the role dropdown reads this; users do not edit it directly
                         "ontology_terms")

# Hidden by default and revealed by the "Show Common Sheets" toggle: the SBH_*
# scaffolds and organism_terms. The user may want to read/extend these, but they
# are not part of routine data entry, so they start hidden.
_HIDDEN_SHEET_PREFIXES = ("SBH_",)
_HIDDEN_COMMON_SHEETS = ("organism_terms",)


def _apply_sheet_visibility(wb):
    for name in _HIDDEN_SYSTEM_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    for name in _HIDDEN_COMMON_SHEETS:
        if name in wb.sheetnames:
            wb[name].sheet_state = "hidden"
    for name in wb.sheetnames:
        if name.startswith(_HIDDEN_SHEET_PREFIXES):
            wb[name].sheet_state = "hidden"


def _apply_reference_styling(wb):
    for sheet_name in ("ontology_terms", "organism_terms", "molecule_types",
                       "nci_thesaurus", "flapjack_cols"):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            _mark_system_sheet(ws)
            ws.freeze_panes = "A2"


def _apply_workbook_col_widths(wb):
    for ws in wb.worksheets:
        _set_fixed_col_widths(ws)


# ── Native Excel checkbox injection ───────────────────────────────────────────
#
# Excel's native checkbox cell control (Insert > Checkbox, 2024+) renders a
# boolean cell as a checkbox. It is stored as a style->feature-bag link that
# openpyxl cannot write (and strips on save), so it is injected by post-processing
# the saved package: a featurePropertyBag part, a content-type override, a
# workbook relationship, a checkbox `xf` in styles.xml, and an `s` attribute on
# each target boolean cell pointing at that xf.

_FPB_PART_PATH    = "xl/featurePropertyBag/featurePropertyBag.xml"
_FPB_CONTENT_TYPE = "application/vnd.ms-excel.featurepropertybag+xml"
_FPB_REL_TYPE     = "http://schemas.microsoft.com/office/2022/11/relationships/FeaturePropertyBag"
_XFPB_NS          = "http://schemas.microsoft.com/office/spreadsheetml/2022/featurepropertybag"
_XFPB_EXT_URI     = "{C7286773-470A-42A8-94C5-96B5CB345126}"
_SS_NS            = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OREL_NS          = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

_FEATURE_PROPERTY_BAG_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    '<FeaturePropertyBags xmlns="%s">'
    '<bag type="Checkbox"/>'
    '<bag type="XFControls"><bagId k="CellControl">0</bagId></bag>'
    '<bag type="XFComplement"><bagId k="XFControls">1</bagId></bag>'
    '<bag type="XFComplements" extRef="XFComplementsMapperExtRef">'
    '<a k="MappedFeaturePropertyBags"><bagId>2</bagId></a></bag>'
    '</FeaturePropertyBags>' % _XFPB_NS
)

# Boolean headers to render as checkboxes
_CHECKBOX_INIT_HEADERS = ("Convert", "Has Collections", "Has Descripts")
# The Translate columns no longer use native checkbox cell-controls: those do not
# propagate on Excel Table expansion. They use a VBA-toggled Unicode ballot box
# instead (see _write_formula_row). Native checkboxes remain only on the fixed
# Init and welcome sheets, which never expand.
_CHECKBOX_DATA_HEADERS: tuple = ()
_NON_DATA_SHEETS = {
    "Init", "column_definitions", "welcome",
}


def _checkbox_targets(out_path: str) -> dict:
    """Return {sheet_name: {(row, col_idx): bool}} of cells to render as checkboxes.

    The boolean is the cell's intended value (read from openpyxl in memory, which
    is reliable regardless of how the value was serialized on disk). The XML pass
    rewrites each target as a proper boolean cell, so a checkbox always renders:
    a native checkbox control only appears on a boolean-typed cell, and openpyxl
    can otherwise emit these as text or empty cells, which is why Update and the
    Translate columns rendered inconsistently.
    """
    wb = openpyxl.load_workbook(out_path,
                                keep_vba=out_path.lower().endswith(".xlsm"))
    targets: dict = {}

    def _as_bool(v):
        if isinstance(v, bool):
            return v
        return str(v).strip().lower() in ("true", "1")

    try:
        if "Init" in wb.sheetnames:
            ws = wb["Init"]
            hdr = {ws.cell(_INIT_HEADER_ROW, c).value: c
                   for c in range(1, ws.max_column + 1)}
            name_col = hdr.get("Sheet Name", 1)
            cols = [hdr[h] for h in _CHECKBOX_INIT_HEADERS if h in hdr]
            cells = {}
            for r in range(_INIT_HEADER_ROW + 1, ws.max_row + 1):
                if ws.cell(r, name_col).value not in (None, ""):
                    for c in cols:
                        cells[(r, c)] = _as_bool(ws.cell(r, c).value)
            if cells:
                targets["Init"] = cells

        if "welcome" in wb.sheetnames:
            ws = wb["welcome"]
            label_row = {str(ws.cell(r, 2).value or "").strip(): r
                         for r in range(1, ws.max_row + 1) if ws.cell(r, 2).value}
            cells = {(label_row[lbl], 3): _as_bool(ws.cell(label_row[lbl], 3).value)
                     for lbl in _WELCOME_CHECKBOX_LABELS if lbl in label_row}
            if cells:
                targets["welcome"] = cells

        for name in wb.sheetnames:
            if name in _NON_DATA_SHEETS:
                continue
            ws = wb[name]
            hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            cells = {(2, hdr[h]): _as_bool(ws.cell(2, hdr[h]).value)
                     for h in _CHECKBOX_DATA_HEADERS if h in hdr}
            if cells:
                targets.setdefault(name, {}).update(cells)
    finally:
        wb.close()
    return targets


def _sheet_xml_map(zin: zipfile.ZipFile) -> dict:
    """Map worksheet display name -> archive path (xl/worksheets/sheetN.xml)."""
    wb_xml   = ET.fromstring(zin.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    rid_to_target = {rel.get("Id"): rel.get("Target") for rel in rels_xml}

    name_to_path = {}
    sheets_el = wb_xml.find("{%s}sheets" % _SS_NS)
    if sheets_el is None:
        return name_to_path
    for sheet in sheets_el:
        rid    = sheet.get("{%s}id" % _OREL_NS)
        target = rid_to_target.get(rid, "")
        if not target:
            continue
        target = ("xl/" + target) if not target.startswith("/") else target.lstrip("/")
        name_to_path[sheet.get("name")] = target
    return name_to_path


def _set_cell_styles(worksheet_bytes: bytes, values: dict, style_index: int) -> bytes:
    """Rewrite each target cell as a checkbox-styled boolean cell.

    `values` maps a cell ref ("N2") to the intended boolean. Each matching cell
    element is replaced whole with `<c r="REF" s="idx" t="b"><v>1|0</v></c>`, so
    the type is always boolean (native checkboxes require it) and the style points
    at the checkbox xf. Replacing the whole element also drops any stale inline
    string or empty body left by the original serialization.
    """
    text = worksheet_bytes.decode("utf-8")
    for ref, val in values.items():
        # match the whole cell element, self-closed or with a body
        pat = re.compile(r'<c\s+r="%s"(?:\s[^>]*)?(?:/>|>.*?</c>)' % re.escape(ref))
        cell = '<c r="%s" s="%d" t="b"><v>%d</v></c>' % (ref, style_index, 1 if val else 0)
        text = pat.sub(lambda m, c=cell: c, text, count=1)
    return text.encode("utf-8")


def _apply_native_checkboxes(out_path: str):
    targets = _checkbox_targets(out_path)
    if not targets:
        return

    with zipfile.ZipFile(out_path, "r") as zin:
        infos = zin.infolist()
        parts = {info.filename: zin.read(info.filename) for info in infos}
        name_to_path = _sheet_xml_map(zin)

    # 1. styles.xml: append a checkbox xf, capture its index. Used by the fixed
    # Init and welcome sheets only (the Translate columns switched to a Unicode
    # ballot box, so no named style is needed here anymore).
    styles = parts["xl/styles.xml"].decode("utf-8")
    m = re.search(r'<cellXfs count="(\d+)">', styles)
    if not m:
        return  # unexpected structure; leave the file untouched
    base_count     = int(m.group(1))
    checkbox_index = base_count
    checkbox_xf = (
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0">'
        '<extLst><ext uri="%s" xmlns:xfpb="%s">'
        '<xfpb:xfComplement i="0"/></ext></extLst></xf>' % (_XFPB_EXT_URI, _XFPB_NS)
    )
    styles = styles.replace('<cellXfs count="%d">' % base_count,
                            '<cellXfs count="%d">' % (base_count + 1), 1)
    styles = styles.replace("</cellXfs>", checkbox_xf + "</cellXfs>", 1)
    parts["xl/styles.xml"] = styles.encode("utf-8")

    # 2. worksheet cells - point each target at the checkbox xf
    for sheet_name, cells in targets.items():
        path = name_to_path.get(sheet_name)
        if not path or path not in parts:
            continue
        values = {f"{get_column_letter(c)}{r}": val for (r, c), val in cells.items()}
        parts[path] = _set_cell_styles(parts[path], values, checkbox_index)

    # 3. featurePropertyBag part
    parts[_FPB_PART_PATH] = _FEATURE_PROPERTY_BAG_XML.encode("utf-8")

    # 4. content types override
    ct = parts["[Content_Types].xml"].decode("utf-8")
    if _FPB_CONTENT_TYPE not in ct:
        override = ('<Override PartName="/%s" ContentType="%s"/>'
                    % (_FPB_PART_PATH, _FPB_CONTENT_TYPE))
        ct = ct.replace("</Types>", override + "</Types>", 1)
        parts["[Content_Types].xml"] = ct.encode("utf-8")

    # 5. workbook relationship
    rels = parts["xl/_rels/workbook.xml.rels"].decode("utf-8")
    if _FPB_REL_TYPE not in rels:
        existing = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels)]
        new_rid  = "rId%d" % ((max(existing) + 1) if existing else 1)
        rel = ('<Relationship Id="%s" Type="%s" '
               'Target="featurePropertyBag/featurePropertyBag.xml"/>'
               % (new_rid, _FPB_REL_TYPE))
        rels = rels.replace("</Relationships>", rel + "</Relationships>", 1)
        parts["xl/_rels/workbook.xml.rels"] = rels.encode("utf-8")

    # 6. rewrite the archive, preserving every original part
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".zip", dir=os.path.dirname(out_path))
    os.close(tmp_fd)
    with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        written = set()
        for info in infos:
            zout.writestr(info, parts[info.filename])
            written.add(info.filename)
        for name, data in parts.items():
            if name not in written:
                zout.writestr(name, data)
    # Windows: os.replace maps to MoveFileEx, which needs DELETE access on the
    # destination and fails with "Access is denied" while any handle to it is
    # open. POSIX rename() has no such restriction, so this only breaks on
    # Windows. Reading the workbook above leaves the source archive's underlying
    # file object alive: ZipFile.close() sets fp to None but only closes the real
    # handle once its internal _fileRefCnt reaches zero, so a member stream that
    # was never closed keeps it open, held in a reference cycle that refcounting
    # cannot reclaim. Collecting the cycles releases it before the swap.
    gc.collect()
    try:
        os.replace(tmp_path, out_path)
    except PermissionError:
        # Fallback if a handle still survives: overwriting in place needs only
        # write access, not DELETE. This gives up the atomicity of the rename,
        # acceptable because out_path is a freshly generated file, and it never
        # leaves the temp archive behind.
        try:
            with open(tmp_path, "rb") as src, open(out_path, "wb") as dst:
                shutil.copyfileobj(src, dst)
        finally:
            os.remove(tmp_path)


# ── Utilities ─────────────────────────────────────────────────────────────────

def _safe_filename(name: str) -> str:
    safe = re.sub(r"[^\w\s\-]", "", name).strip()
    safe = re.sub(r"\s+", "_", safe)
    return safe or "MyLibrary"


def _type_label(template_type: str, config: dict) -> str:
    labels = {
        "resources":     "Resources",
        "strains":       "Strains",
        "sample_design": "SampleDesign",
        "study":         "Study",
    }
    if template_type == "custom":
        return _safe_filename(config.get("template_name", "Custom"))
    return labels.get(template_type, template_type.title())
