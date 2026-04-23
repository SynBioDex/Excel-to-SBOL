#Validator code
from __future__ import annotations

from dataclasses import dataclass
from email.mime import message
from typing import Any, Dict, List, Optional

import os

# We reuse the existing compiler in 1.1.18 so we don't reinvent workbook parsing yet.
# Later you'll move checks into compiler.py, but for now this keeps the validator runnable quickly.
from urllib.parse import urlparse
import excel2sbol.compiler as compiler
import pandas as pd
from openpyxl import load_workbook

def _df_cellmap(df, fn):
    """
    Pandas compatibility:
    - Newer pandas: DataFrame.map exists (elementwise)
    - Older pandas: use DataFrame.applymap
    """
    if hasattr(df, "map"):
        return df.map(fn)
    return df.applymap(fn)

# -----------------------------
# Validation data structures
# -----------------------------

@dataclass
class ValidationItem:
    sheet: str
    row_display_id: Optional[str]   # for sheet-level checks this is usually None
    column: Optional[str]
    code: str
    message: str


class ValidationCollector:
    """
    validate_only=True: collect errors/warnings and keep going.
    validate_only=False: raise on first error (fail-fast).
    echo=True: print each error/warning as it's added.
    verbose=True: return the full validation payload.
    verbose=False: return a minimal payload.
    """
    def __init__(self, validate_only: bool = True, echo: bool = False, verbose: bool = True):
        self.validate_only = validate_only
        self.echo = echo
        self.verbose = verbose
        self.errors: List[ValidationItem] = []
        self.warnings: List[ValidationItem] = []

    def build_payload(self, ok: bool, validated_sheets: Optional[List[str]] = None) -> Any:
        if self.verbose:
            return {
                "ok": ok,
                "errors": [e.__dict__ for e in self.errors],
                "warnings": [w.__dict__ for w in self.warnings],
                "validated_sheets": validated_sheets or [],
            }
        return "Pass" if ok else None

    def error(self, sheet: str, row_display_id: Optional[str], column: Optional[str], code: str, message: str) -> None:
    # Force one-line messages (collapse whitespace/newlines)
        msg = " ".join(str(message).split())
        item = ValidationItem(sheet=sheet, row_display_id=row_display_id, column=column, code=code, message=msg)
        self.errors.append(item)
        if self.echo:
            print(f"[ERROR] ({code}) {sheet}/{column or '-'}: {msg}")
        if not self.validate_only:
            raise ValueError(f"[{code}] {sheet}/{column or '-'}: {msg}")

    def warn(self, sheet: str, row_display_id: Optional[str], column: Optional[str], code: str, message: str) -> None:
        # Force one-line messages (collapse whitespace/newlines)
        msg = " ".join(str(message).split())
        item = ValidationItem(sheet=sheet, row_display_id=row_display_id, column=column, code=code, message=msg)
        self.warnings.append(item)
        if self.echo:
            print(f"[WARN]  ({code}) {sheet}/{column or '-'}: {msg}")
            
def _msg(code: str, sheet: str, column: Optional[str], detail: str, fix: Optional[str] = None) -> str:
    """
    Standard one-line message builder.
    Format: CODE: sheet/column — detail. Fix: ...
    """
    base = f"{code}: {sheet}/{column or '-'} — {detail}"
    return base if not fix else f"{base}. Fix: {fix}"        
        
# -----------------------------
# Baby checks (add more later)
# -----------------------------

def _read_init_sheet_names(file_path_in: str) -> List[str]:
    """
    Read the Init table (same shape compiler.initialise expects) and return the sheet names listed there.
    In 1.1.18, Init is read with skiprows=9 and index_col=0.
    """
    init_df = pd.read_excel(
        file_path_in,
        sheet_name="Init",
        skiprows=9,
        index_col=0,
        engine="openpyxl",
    )
    # strip whitespace from all string cells (pandas 2.x/3.x safe)
    init_df = _df_cellmap(init_df, lambda x: x.strip() if isinstance(x, str) else x)
    # the index values are the sheet names
    return [str(x).strip() for x in init_df.index.tolist() if str(x).strip() != ""]


def _read_workbook_sheetnames(file_path_in: str) -> List[str]:
    wb = load_workbook(file_path_in, data_only=True, read_only=True, keep_links=False)
    return list(wb.sheetnames)


def _check_sheet_names_in_init_and_exist(
    col_read_df: pd.DataFrame,
    init_sheet_names: List[str],
    workbook_sheetnames: List[str],
    validator: ValidationCollector,
) -> None:
    """
    SHEET_NAME_* checks (from column_definitions):
      - Sheet Name must be listed in Init
      - Sheet Name must exist as an actual worksheet
    """
    init_set = set(init_sheet_names)
    wb_set = set(workbook_sheetnames)

    # column_definitions can contain NaN; normalize carefully
    unique_sheets = []
    for v in col_read_df.get("Sheet Name", []).dropna().unique():
        s = str(v).strip()
        if not s:
            continue
        # ignore these definition/control sheets
        if s.lower() in {"init", "column_definitions"}:
            continue
        unique_sheets.append(s)

    for sht in sorted(set(unique_sheets)):
        if sht not in init_set:
            validator.error(
                sheet=sht,
                row_display_id=None,
                column="Sheet Name",
                code="SHEET_NOT_IN_INIT",
                message=_msg(
                    "SHEET_NOT_IN_INIT",
                    sht,
                    "Sheet Name",
                    'Sheet appears in column_definitions but is not listed in Init',
                    fix='Add sheet to Init or correct Sheet Name in column_definitions',
                ),
            )

        if sht not in wb_set:
            validator.warn(
                sheet=sht,
                row_display_id=None,
                column="Sheet Name",
                code="MISSING_SHEET",
                message=_msg(
                    "MISSING_SHEET",
                    sht,
                    "Sheet Name",
                    "Sheet appears in column_definitions but does not exist in the workbook",
                    fix="Create the worksheet or correct the Sheet Name",
                ),
            )

def _warn_extra_sheet_columns(col_read_df, compiled_sheets, to_convert, validator: ValidationCollector) -> None:
    """
    WARNING: Column exists in sheet but is NOT declared in column_definitions.
    """
    for sht in to_convert:
        lib = compiled_sheets.get(sht, {}).get("library", {})
        if not lib:
            continue

        for col in lib.keys():
            col_norm = str(col).strip()
            # ignore helper columns that may exist without being declared
            if col_norm.lower() in {"update", "uri"}:
                continue

            match = col_read_df.loc[
                (col_read_df["Sheet Name"] == sht) &
                (col_read_df["Column Name"] == col_norm)
            ]
            if match.empty:
                validator.warn(
                    sheet=sht,
                    row_display_id=None,
                    column=col_norm,
                    code="UNDECLARED_COLUMN",
                    message=_msg(
                        "UNDECLARED_COLUMN",
                        sht,
                        col_norm,
                        "Column exists in sheet but is missing from column_definitions (extra/unexpected column)",
                        fix="Add it to column_definitions or rename/remove the column",
                    ),
                )

def _error_missing_sheet_columns(col_read_df, compiled_sheets, to_convert, validator: ValidationCollector) -> None:
    """
    ERROR: Column is declared in column_definitions for a sheet but is NOT present in the sheet data.
    """
    to_convert_set = set(to_convert)

    for sht in to_convert:
        lib = compiled_sheets.get(sht, {}).get("library", {})
        if not lib:
            continue
        lib_cols = set(str(c).strip() for c in lib.keys())

        # all declared columns for this sheet
        defs = col_read_df.loc[col_read_df["Sheet Name"] == sht]
        if defs.empty:
            continue

        for v in defs["Column Name"].dropna().tolist():
            colname = str(v).strip()
            if not colname:
                continue
            # ignore helper columns
            if colname.lower() in {"update", "uri"}:
                continue

            if colname not in lib_cols:
                validator.error(
                    sheet=sht,
                    row_display_id=None,
                    column=colname,
                    code="COLUMN_DEF_MISSING_IN_SHEET",
                    message=_msg(
                        "COLUMN_DEF_MISSING_IN_SHEET",
                        sht,
                        colname,
                        "Column is declared in column_definitions but missing from the sheet header",
                        fix="Add the column to the sheet or remove/fix the row in column_definitions",
                    ),
                )
                
def _is_blank(x) -> bool:
    return x is None or (isinstance(x, float) and pd.isna(x)) or str(x).strip() == ""


def _is_not_applicable(x) -> bool:
    return (not _is_blank(x)) and str(x).strip().lower() == "not_applicable"


def _is_valid_url(s: str) -> bool:
    try:
        u = urlparse(s.strip())
        return u.scheme in {"http", "https"} and bool(u.netloc)
    except Exception:
        return False


def _split_on_makes_sense(split_on_val) -> bool:
    """
    Valid Split On:
      - MUST be present (blank cell is NOT ok)
      - MUST be a quoted string, including empty quotes: "" is allowed
        Examples valid:  ""   "."   ","   " | "
        Examples invalid: (blank)   .   ,   "   " (missing closing)   abc
    """
    if _is_blank(split_on_val):
        return False  # blank cell not ok

    s = str(split_on_val).strip()
    return len(s) >= 2 and s.startswith('"') and s.endswith('"')

    # Otherwise accept a simple delimiter token (commas, semicolons, pipes, whitespace, etc.)
    return True


def _check_column_def_sbol_term_and_related_fields(col_read_df: pd.DataFrame, validator: ValidationCollector) -> None:
    """
    Check rules (per supervisor request):
    - SBOL Term must be filled.
      EXCEPTION: if SBOL Term is 'not_applicable' (or Type is 'not_applicable'), skip the rest;
                 SBOL Term can be empty in that case.
    - If not not_applicable:
        - Namespace URL must be a valid URL (http/https)
        - Type must be filled
        - Split On must be filled with something that makes sense
    """

    required_cols = {"Sheet Name", "Column Name", "SBOL Term", "Namespace URL", "Type", "Split On"}
    missing_cols = [c for c in required_cols if c not in col_read_df.columns]
    if missing_cols:
        validator.error(
            sheet="column_definitions",
            row_display_id=None,
            column=None,
            code="COLUMN_DEFS_MALFORMED",
            message=_msg(
                "COLUMN_DEFS_MALFORMED",
                "column_definitions",
                None,
                f"Missing required columns: {missing_cols}",
                fix="Add these header columns to column_definitions",
            ),
        )
        return

    for _, row in col_read_df.iterrows():
        sheet_name = str(row.get("Sheet Name", "")).strip()
        col_name = str(row.get("Column Name", "")).strip()

        # Skip rows that are obviously not real defs
        if sheet_name == "" or col_name == "":
            continue
        if sheet_name.lower() in {"init"}:
            continue

        sbol_term = row.get("SBOL Term", None)
        ns_url = row.get("Namespace URL", None)
        typ = row.get("Type", None)
        split_on = row.get("Split On", None)

        # Determine whether this row is "not_applicable"
        is_na = _is_not_applicable(sbol_term) or _is_not_applicable(typ)

        # Rule: SBOL Term must be filled unless not_applicable
        if not is_na and _is_blank(sbol_term):
            validator.error(
                sheet=sheet_name,
                row_display_id=None,
                column=col_name,
                code="SBOL_TERM_MISSING",
                message=_msg(
                    "SBOL_TERM_MISSING",
                    sheet_name,
                    col_name,
                    "SBOL Term is empty",
                    fix='Fill SBOL Term or set it to "not_applicable"',
                ),
            )
            # can't validate the rest sensibly without an SBOL Term
            continue

        # If not_applicable: skip all other checks (and allow blanks)
        if is_na:
            continue

        # Namespace URL must be filled + valid URL
        if _is_blank(ns_url) or not _is_valid_url(str(ns_url)):
            validator.error(
                sheet=sheet_name,
                row_display_id=None,
                column=col_name,
                code="NAMESPACE_URL_INVALID",
                message=_msg(
                    "NAMESPACE_URL_INVALID",
                    sheet_name,
                    col_name,
                    f'Namespace URL must be a valid http/https URL; got "{"" if _is_blank(ns_url) else str(ns_url).strip()}"',
                    fix="Enter a full URL like https://example.org",
                ),
            )

        # Type must be filled
        if _is_blank(typ):
            validator.error(
                sheet=sheet_name,
                row_display_id=None,
                column=col_name,
                code="TYPE_MISSING",
                message=_msg(
                    "TYPE_MISSING",
                    sheet_name,
                    col_name,
                    "Type is empty",
                    fix="Fill Type or set it to not_applicable",
                ),
            )

        # Split On must be filled with something that makes sense
        if not _split_on_makes_sense(split_on):
            validator.error(
                sheet=sheet_name,
                row_display_id=None,
                column=col_name,
                code="SPLIT_ON_INVALID",
                message=_msg(
                    "SPLIT_ON_INVALID",
                    sheet_name,
                    col_name,
                    "Split On is invalid (must be a quoted delimiter like \"\" or \",\")",
                    fix='Set Split On to "" for no-split, or to "," for comma-separated lists',
                ),
            )
            
def _check_lookup_sheet_exists(col_read_df: pd.DataFrame,
                               workbook_sheetnames: List[str],
                               validator: ValidationCollector) -> None:
    """
    Lookup Sheet: if column_definitions specifies a Lookup Sheet, that sheet must exist in the workbook.
    """
    if "Lookup Sheet Name" not in col_read_df.columns:
        # If the template doesn't use this feature, don't fail.
        validator.warn(
            sheet="column_definitions",
            row_display_id=None,
            column="Lookup Sheet Name",
            code="LOOKUP_SHEET_COLUMN_MISSING",
            message=_msg(
                "LOOKUP_SHEET_COLUMN_MISSING",
                "column_definitions",
                "Lookup Sheet Name",
                'No "Lookup Sheet Name" column; skipping Lookup Sheet validation',
            ),
        )
        return

    wb_set = set(workbook_sheetnames)

    for _, row in col_read_df.iterrows():
        sheet_name = str(row.get("Sheet Name", "")).strip()
        col_name = str(row.get("Column Name", "")).strip()

        # ignore non-real rows
        if sheet_name == "" or col_name == "":
            continue
        if sheet_name.lower() in {"init"}:
            continue

        lookup = row.get("Lookup Sheet Name", None)
        if _is_blank(lookup):
            continue

        lookup_name = str(lookup).strip()
        if lookup_name == "":
            continue

        if lookup_name not in wb_set:
            validator.error(
                sheet=sheet_name,
                row_display_id=None,
                column=col_name,
                code="LOOKUP_SHEET_MISSING",
                message=_msg(
                    "LOOKUP_SHEET_MISSING",
                    sheet_name,
                    col_name,
                    f'Lookup Sheet Name "{lookup_name}" does not exist in the workbook',
                    fix="Create the lookup sheet or correct Lookup Sheet Name",
                ),
            )
            
def _check_init_vs_workbook_sheets(
    init_sheet_names: List[str],
    workbook_sheetnames: List[str],
    validator: ValidationCollector,
) -> None:
    """
    Supervisor rule:
      - ERROR if a sheet is declared in Init but does not exist as a tab in the workbook
      - WARNING if a tab exists in the workbook but is not declared in Init
    """

    init_set = set(s.strip() for s in init_sheet_names if str(s).strip())
    wb_set = set(s.strip() for s in workbook_sheetnames if str(s).strip())

    # Sheets you do NOT want to enforce/warn about (customize freely)
    exclude_from_extra = {
        "Init",
        "column_definitions",
        "protein_and_complex_Ids"
    }
    exclude_from_required = {
        # put sheets here ONLY if you do NOT want missing-Init-tab to error
        # e.g. "welcome"
    }

    # ERROR: Init declares it, but workbook tab missing
    for s in sorted(init_set):
        if s in exclude_from_required:
            continue
        if s not in wb_set:
            validator.error(
                sheet="Init",
                row_display_id=None,
                column="Sheet Name",
                code="INIT_SHEET_MISSING",
                message=_msg(
                    "INIT_SHEET_MISSING",
                    "Init",
                    "Sheet Name",
                    f'Init declares sheet "{s}" but workbook has no tab with that name',
                    fix="Create the worksheet tab or remove/fix the entry in Init",
                ),
            )

    # WARNING: workbook has a tab not declared in Init
    for s in sorted(wb_set):
        if s in exclude_from_extra:
            continue
        if s not in init_set:
            validator.warn(
                sheet="Init",
                row_display_id=None,
                column="Sheet Name",
                code="WORKBOOK_SHEET_NOT_IN_INIT",
                message=_msg(
                    "WORKBOOK_SHEET_NOT_IN_INIT",
                    "Init",
                    "Sheet Name",
                    f'Workbook contains sheet "{s}" but it is not declared in Init',
                    fix="Add it to Init or delete/rename the worksheet tab",
                ),
            )

# -----------------------------
# Big runner function
# -----------------------------

def run_sheet_validator(
    file_path_in: str,
    *,
    validate_only: bool = True,
    echo: bool = False,
    verbose: bool = True,
) -> Any:
    """
    Validate a workbook using sheet-level checks.

    Now includes:
      - SHEET_NOT_IN_INIT
      - MISSING_SHEET
      - UNDECLARED_COLUMN
    """
    validator = ValidationCollector(validate_only=validate_only, echo=echo, verbose=verbose)

    # Read column_definitions directly (so we can validate sheet names even if compiler.initialise would crash)
    col_read_df = pd.read_excel(
        file_path_in,
        sheet_name="column_definitions",
        header=0,
        engine="openpyxl",
    )
    col_read_df = _df_cellmap(col_read_df,lambda x: x.strip() if isinstance(x, str) else x)

    # --- NEW baby check: Sheet Name must be in Init and must exist in workbook ---
    init_sheet_names = _read_init_sheet_names(file_path_in)
    workbook_sheetnames = _read_workbook_sheetnames(file_path_in)
    _check_init_vs_workbook_sheets(init_sheet_names, workbook_sheetnames, validator)
    _check_sheet_names_in_init_and_exist(col_read_df, init_sheet_names, workbook_sheetnames, validator)
    _check_lookup_sheet_exists(col_read_df, workbook_sheetnames, validator)
    _check_column_def_sbol_term_and_related_fields(col_read_df, validator)

    # If we're collecting (not fail-fast) and we already know sheets are missing/misdeclared,
    # stop early to avoid downstream crashes inside compiler.initialise.
    if validate_only and len(validator.errors) > 0:
        ok = False
        return validator.build_payload(ok=ok, validated_sheets=[])

    # Reuse 1.1.18's compiler parsing for checks that need compiled_sheets/to_convert
    col_read_df2, to_convert, compiled_sheets, _version_info, _homespace = compiler.initialise(file_path_in)

    # Keep using the compiler-produced col_read_df for downstream checks (it may have extra processing)
    _warn_extra_sheet_columns(col_read_df2, compiled_sheets, to_convert, validator)
    _error_missing_sheet_columns(col_read_df2, compiled_sheets, to_convert, validator)

    ok = (len(validator.errors) == 0)
    return validator.build_payload(ok=ok, validated_sheets=list(to_convert))


# -----------------------------
# Optional CLI entrypoint
# -----------------------------

def _main() -> None:
    import argparse
    import json

    p = argparse.ArgumentParser(description="Validate an excel2sbol workbook (sheet-level checks).")
    p.add_argument("input", help="Path to input .xlsx/.xlsm workbook")
    p.add_argument("--echo", action="store_true", help="Print errors/warnings as they are found")
    p.add_argument("--fail-fast", action="store_true", help="Raise on first error instead of collecting")
    args = p.parse_args()

    result = run_sheet_validator(args.input, validate_only=(not args.fail_fast), echo=args.echo)
    print(json.dumps(result, indent=2))

    # conventional exit code: 0 ok, 1 errors found
    raise SystemExit(0 if result["ok"] else 1)


if __name__ == "__main__":
    _main()
