import re
import string
import rdflib
from openpyxl.worksheet import cell_range, worksheet
from openpyxl import load_workbook
from pathlib import Path

def check_name(nm_to_chck):
    """the function verifies that the names is alphanumeric and
    separated by underscores if that is not the case the special characters are
    replaced by their unicode decimal code number

    Args:
        nm_to_chck (string): the name to be checked

    Returns:
        compliant_name (string): alphanumberic name with special
                                 characters replaced by _u###_
    """

    if not bool(re.match('^[a-zA-Z0-9]+$', nm_to_chck)):
        # replace special characters with numbers
        for ltr in nm_to_chck:
            if ord(ltr) == 32:
                nm_to_chck = nm_to_chck.replace(ltr, "_")
            elif ord(ltr) == 45:
                # Allow hyphens to be reinterpreted as underscores
                nm_to_chck = nm_to_chck.replace(ltr, "_")
            elif ord(ltr) == 46:
                nm_to_chck = nm_to_chck.replace(ltr, "_")
            elif ord(ltr) > 122 or ord(ltr) < 48:
                # 122 is the highest decimal code number
                # for common latin ltrs or arabic numbers
                # this helps identify special characters like
                # ä or ñ, which isalnum() returns as true
                # the characters that don't meet this criterion are replaced
                # by their decimal code number separated by an underscore
                nm_to_chck = nm_to_chck.replace(ltr, str(f"_u{ord(ltr)}_"))
                # new_ltr = str(ltr.encode("unicode_escape"))
                # new_ltr = new_ltr.replace(r"b'\\", "").replace("'", "")
                # nm_to_chck = nm_to_chck.replace(ltr, f'_{new_ltr}_')
            else:
                # remove all letters, numbers and whitespaces
                ltr = re.sub(r'[\w, \s]', '', ltr)
                # this enables replacing all other
                # special characters that are under 122
                if len(ltr) > 0:
                    nm_to_chck = nm_to_chck.replace(ltr, str(f"_u{ord(ltr)}_"))
    if len(nm_to_chck) > 0:
        if nm_to_chck[0].isnumeric():
            # ensures it doesn't start with a number
            nm_to_chck = f"_{nm_to_chck}"

    return(nm_to_chck)


def truthy_strings(to_check):
    """Takes in several variants of True and False and returns a boolean
    True or False

    Args:
        to_check (string,boolean): A string or boolean such as 'True', 'TRUE',
        'tRue', or True

    Raises:
        TypeError: If the value can't be converted to 'true' or 'false'
        an error is raised

    Returns:
        [boolean]: True or False is returned depending on the inputs
    """
    if str(to_check).lower() == 'false':
        return False
    elif str(to_check).lower() == 'true':
        return True
    else:
        raise TypeError


def col_to_num(col_name):
    """takes an excel column name, e.g. AA and converts it to a
    one indexed number e.g. 27

    Args:
       col_name (string): An excel formatted column name, e.g. AA

    Raises:
        TypeError: Raised if the input is not a string
        ValueError: Raised if the string is longer than three
                    or contains spaces

    Returns:
        num (integer): A one indexed column index
    """

    if type(col_name) != str:
        # is not a string
        raise TypeError
    elif len(col_name.replace(" ", "")) != len(col_name):
        # contains spaces
        raise ValueError
    elif len(col_name) > 3:
        # too long to be an excel column name
        raise ValueError
    return cell_range.range_boundaries(f'{col_name}1')[0]

def row_ends(sheet: worksheet, row: int, min_col: int = 1) -> int:
    """Find the column at which a row ends
    :param sheet: Sheet to search
    :param row: Row to search
    :param min_col: Column on which to start, defaults to first column
    :return: numerical value of last column containing data
    """
    # convert min column to numerical to determine starting point
    column_iterator = sheet.iter_cols(min_row=row, max_row=row, min_col=min_col, values_only=True)
    max_sequence_length = 100000  # TODO: use a proper generator counter rather than a big range
    return next(i-1 for v, i in zip(column_iterator, range(min_col, max_sequence_length)) if not v[0])


def read_variant_table(excel_file: Path) -> tuple[str, str, list[list]]:
    """Extract an amino acid site-variant table from a Twist amino-acid site-variant library Excel sheet
    :param excel_file: location of file to read
    :return: Tuple of library name, original sequence, list of variant lists for each site, in order
    """
    VARIANTS_SHEET = 'Amino Acid Variants'
    LIBRARY_NAME_CELL = 'C10'
    FIRST_AMINO_ACID_COLUMN = 'F'
    ORIGINAL_AMINO_ACID_ROW = 12
    FIRST_VARIANT_ROW = 14
    LAST_VARIANT_ROW = 35
    
    print(f'Loading workbook "{excel_file}"')
    work_book = load_workbook(excel_file, data_only=True)
    sheet = work_book[VARIANTS_SHEET]

    # First, get the library name
    library_name = sheet[LIBRARY_NAME_CELL].value
    print(f'Library is named "{library_name}"')

    # Then get the base sequence
    print('Extracting base sequence')
    first_aa_column = col_to_num(FIRST_AMINO_ACID_COLUMN)
    last_aa_column = row_ends(sheet, ORIGINAL_AMINO_ACID_ROW, first_aa_column)
    # Get row from sheet and concatenate it into a string
    row_iterator = sheet.iter_rows(min_row=ORIGINAL_AMINO_ACID_ROW, max_row=ORIGINAL_AMINO_ACID_ROW,
                                   min_col=first_aa_column, max_col=last_aa_column, values_only=True)
    base_sequence = [''.join(row) for row in row_iterator][0]
    print(f'Found sequence {len(base_sequence)} residues long: "{base_sequence}"')

    # Finally, get all the variant lists
    column_iterator = sheet.iter_cols(min_row=FIRST_VARIANT_ROW, max_row=LAST_VARIANT_ROW,
                                      min_col=first_aa_column, max_col=last_aa_column, values_only=True)
    variant_lists = [[v for v in column if v] for column in column_iterator]  # drop the empty cells from each range

    return library_name, base_sequence, variant_lists

def update_uri_refs(doc, update_dict, use_derived=True, derived_ls = ['_sequence']):
    """
    This updates a set of referenced uris (may be a namespace or identity update)

    Args:
        doc (SBOL3 Document): document to be updated
        update_dict (dict): dictionary of the form {old_uri:new_uri}
        use_derived (bool, optional): Whether or not to also update derived uris. Defaults to True.
        derived_ls (list, optional): List of derivations e.g. also version of the uri
                                     with _sequence added to the end. Defaults to ['_sequence'].

    Returns:
        doc (SBOL3 Document): updated document
    """
    # create all the additional uris that will need to be updated
    derived_keys = []
    for deriv in derived_ls:
        der_update = [f'{x}{deriv}' for x in  update_dict.keys()]
        derived_keys.extend(der_update)

    # pull the graph from the document
    g = doc.graph()
    for index, (subject, predicate, _object) in enumerate(g):
        # if the object is one of the items to be updated do so
        if str(_object) in update_dict:
            g.remove((subject, predicate, _object))
            new = rdflib.URIRef(update_dict[str(_object)])
            g.add((subject, predicate, new))
        # update any derived objects
        elif use_derived and str(_object) in derived_keys:
            suffix = str(_object).split('_')[-1]  # assumes suffix starts with '_'
            suffix = f'_{suffix}'
            g.remove((subject, predicate, _object))
            old = str(_object)
            new = f"{update_dict[old.replace(suffix, '')]}{suffix}"
            new = rdflib.URIRef(new)
            g.add((subject, predicate, new))
        # update any derived subjects
        if use_derived and str(subject) in derived_keys:
            suffix = str(subject).split('_')[-1]  # assumes suffix starts with '_'
            suffix = f'_{suffix}'
            g.remove((subject, predicate, _object))
            old = str(subject)
            new = f"{update_dict[old.replace(suffix, '')]}{suffix}"
            new = rdflib.URIRef(new)
            g.add((new, predicate, _object))
    doc._parse_graph(g)
    return doc